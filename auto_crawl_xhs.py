#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
悟昕小红书数据自动采集脚本（v2 - 双关键词双Sheet）
- Sheet 1: "悟昕"品牌相关笔记
- Sheet 2: 睡眠类低粉高赞爆款笔记（点赞>1000，7个关键词）
- 每周通过 automation 自动触发
"""
import json
import os
import sys
import subprocess
import glob
import time
from datetime import datetime

# === 配置 ===
PROJECT_DIR = "/Users/echochen/MediaCrawler"
OUTPUT_DIR = "/Users/echochen/Desktop/悟昕/03. 数据分析"
PLATFORM = "xhs"
PYTHON_BIN = os.path.join(PROJECT_DIR, ".venv", "bin", "python")

# 任务1: 品牌关键词
BRAND_KEYWORD = "悟昕"

# 任务2: 睡眠爆款关键词
SLEEP_KEYWORDS = "睡眠,失眠,助眠,睡眠仪,入睡困难,深度睡眠,睡眠管理"
HOT_LIKE_THRESHOLD = 1000  # 点赞数阈值

# 品牌官方号排除列表（小写匹配）
BRAND_ACCOUNTS = {"悟昕科技zenoasis", "悟昕", "zenoasis"}


def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def check_singleton():
    """确保没有其他实例在运行"""
    pgrep = subprocess.run(["pgrep", "-f", "MediaCrawler/main.py"], capture_output=True, text=True)
    if pgrep.stdout.strip():
        log(f"警告: 已有MediaCrawler在运行 (PID: {pgrep.stdout.strip()})，退出等待下次执行")
        sys.exit(0)


def run_crawler(keywords, max_notes=60):
    """运行 MediaCrawler 搜索采集"""
    log(f"开始采集，关键词: {keywords}，数量上限: {max_notes}")
    cmd = [
        PYTHON_BIN,
        os.path.join(PROJECT_DIR, "main.py"),
        "--platform", PLATFORM,
        "--type", "search",
        "--keywords", keywords,
        "--save_data_option", "json",
        "--lt", "cookie",
    ]
    result = subprocess.run(
        cmd,
        cwd=PROJECT_DIR,
        capture_output=True,
        text=True,
        timeout=900,  # 15分钟超时（多关键词需要更长时间）
    )
    if result.returncode != 0:
        log(f"采集失败! stderr: {result.stderr[-500:]}")
        log(f"stdout: {result.stdout[-500:]}")
        return None
    log("采集完成")
    return find_latest_json()


def find_latest_json():
    """找到最新的采集结果JSON"""
    pattern = os.path.join(PROJECT_DIR, "data", PLATFORM, "json", "search_contents_*.json")
    files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    if not files:
        log("未找到采集结果JSON文件!")
        return None
    latest = files[0]
    log(f"最新JSON: {latest}")
    return latest


def load_notes(json_path):
    """读取JSON笔记数据"""
    with open(json_path, 'r', encoding='utf-8') as f:
        notes = json.load(f)
    log(f"加载 {len(notes)} 条笔记")
    return notes


def normalize_note(note, seq_num, source_kw=""):
    """统一格式化一条笔记"""
    ts = note.get('time', 0)
    date_str = "未知"
    if ts:
        ts = ts / 1000 if ts > 1000000000000 else ts
        try:
            dt = datetime.fromtimestamp(ts)
            date_str = dt.strftime('%Y-%m-%d %H:%M')
        except Exception:
            pass

    tags = note.get('tag_list', '')
    if isinstance(tags, list):
        tags = ', '.join(tags)

    liked = int(note.get('liked_count', 0) or 0)
    collected = int(note.get('collected_count', 0) or 0)
    comment = int(note.get('comment_count', 0) or 0)
    share = int(note.get('share_count', 0) or 0)

    return {
        '序号': seq_num,
        '标题': note.get('title', '') or '(无标题)',
        '作者': note.get('nickname', ''),
        '发布时间': date_str,
        '点赞数': liked,
        '收藏数': collected,
        '评论数': comment,
        '转发数': share,
        '总互动': liked + collected + comment + share,
        '标签': tags,
        'IP属地': note.get('ip_location', ''),
        '来源关键词': source_kw or note.get('source_keyword', ''),
        '笔记链接': f"https://www.xiaohongshu.com/explore/{note.get('note_id', '')}",
        '笔记ID': note.get('note_id', ''),
    }


def filter_hot_notes(notes):
    """筛选爆款笔记: 点赞>阈值 + 排除品牌号"""
    seen_ids = set()
    filtered = []

    for note in notes:
        liked = int(note.get('liked_count', 0) or 0)
        nickname = (note.get('nickname', '') or '').lower().strip()
        note_id = note.get('note_id', '')

        # 去重
        if note_id in seen_ids:
            continue

        # 点赞阈值
        if liked < HOT_LIKE_THRESHOLD:
            continue

        # 排除品牌官方号
        is_brand = any(brand in nickname for brand in BRAND_ACCOUNTS)
        if is_brand:
            continue

        seen_ids.add(note_id)
        filtered.append(note)

    log(f"爆款筛选: {len(notes)}条 → {len(filtered)}条 (点赞>{HOT_LIKE_THRESHOLD}，已排除品牌号)")
    return filtered


def write_sheet(wb, sheet_name, data, header_color="7C5DFA"):
    """将数据写入 Excel Sheet"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    if not data:
        ws = wb.create_sheet(title=sheet_name)
        ws.cell(row=1, column=1, value="无数据")
        return

    ws = wb.create_sheet(title=sheet_name)

    headers = list(data[0].keys())
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
            cell.alignment = Alignment(vertical="center", wrap_text=(key == "标题"))

    col_widths = {
        '序号': 6, '标题': 45, '作者': 15, '发布时间': 18,
        '点赞数': 10, '收藏数': 10, '评论数': 10, '转发数': 10,
        '总互动': 10, '标签': 30, 'IP属地': 10, '来源关键词': 12,
        '笔记链接': 50, '笔记ID': 25,
    }
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_widths.get(header, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main():
    log("=" * 60)
    log("悟昕小红书数据采集任务启动 (v2 双Sheet)")
    log("=" * 60)

    if not os.path.exists(PYTHON_BIN):
        log(f"错误: Python虚拟环境不存在: {PYTHON_BIN}")
        sys.exit(1)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    check_singleton()

    try:
        import openpyxl
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
        import openpyxl

    wb = openpyxl.Workbook()
    # 删除默认Sheet
    wb.remove(wb.active)

    # ====== 任务1: 采集"悟昕"品牌笔记 ======
    log(">>> 任务1: 采集'悟昕'品牌笔记")
    json1 = run_crawler(BRAND_KEYWORD, max_notes=20)
    if json1:
        notes1 = load_notes(json1)
        notes1_sorted = sorted(notes1, key=lambda x: int(x.get('liked_count', 0) or 0), reverse=True)
        data1 = [normalize_note(n, i + 1) for i, n in enumerate(notes1_sorted)]
        write_sheet(wb, "悟昕品牌笔记", data1, header_color="7C5DFA")
        log(f"Sheet1 写入 {len(data1)} 条")
    else:
        write_sheet(wb, "悟昕品牌笔记", [], header_color="7C5DFA")
        log("Sheet1 采集失败，写入空表")

    # 等待一下，避免连续请求太快
    time.sleep(5)

    # ====== 任务2: 采集睡眠爆款笔记 ======
    log(">>> 任务2: 采集睡眠爆款笔记（7个关键词）")
    json2 = run_crawler(SLEEP_KEYWORDS, max_notes=60)
    if json2:
        notes2 = load_notes(json2)
        # 按关键词搜索后，需要手动读取所有相关JSON（可能多次搜索）
        # MediaCrawler 多关键词时会合并到一个JSON中
        hot_notes = filter_hot_notes(notes2)
        hot_sorted = sorted(hot_notes, key=lambda x: int(x.get('liked_count', 0) or 0), reverse=True)
        data2 = [normalize_note(n, i + 1) for i, n in enumerate(hot_sorted)]
        write_sheet(wb, "睡眠爆款笔记", data2, header_color="E8654A")
        log(f"Sheet2 写入 {len(data2)} 条")
    else:
        write_sheet(wb, "睡眠爆款笔记", [], header_color="E8654A")
        log("Sheet2 采集失败，写入空表")

    # ====== 保存 ======
    today_str = datetime.now().strftime("%Y%m%d")
    output_filename = f"悟昕小红书笔记数据_MediaCrawler_{today_str}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, output_filename)

    if os.path.exists(output_path):
        os.remove(output_path)

    wb.save(output_path)
    log(f"Excel已保存: {output_path}")
    log("=" * 60)
    log("任务完成!")
    log("=" * 60)


if __name__ == "__main__":
    main()
