#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
合并小红书 + last30days 数据为三 Sheet Excel
- Sheet 1: 悟昕品牌笔记
- Sheet 2: 睡眠爆款笔记
- Sheet 3: 全球趋势
"""
import json
import os
import sys
import glob
import re
from datetime import datetime

PROJECT_DIR = "/Users/echochen/MediaCrawler"
OUTPUT_DIR = "/Users/echochen/Desktop/悟昕/03. 数据分析"
HOT_LIKE_THRESHOLD = 100
BRAND_ACCOUNTS = {"悟昕科技zenoasis", "悟昕", "zenoasis"}

# Excel 非法字符正则
ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

def sanitize(text):
    """清除 Excel 不支持的非法控制字符"""
    if not isinstance(text, str):
        return text
    return ILLEGAL_CHARS_RE.sub('', text)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill


def normalize_note(note, seq_num, source_kw=""):
    ts = note.get('time', 0)
    date_str = "未知"
    if ts:
        ts = ts / 1000 if ts > 1000000000000 else ts
        try:
            dt = datetime.fromtimestamp(ts)
            date_str = dt.strftime('%Y-%m-%d %H:%M')
        except Exception:
            pass

    tags = note.get('tag_list', '')
    if isinstance(tags, list):
        tags = ', '.join(tags)

    liked = int(note.get('liked_count', 0) or 0)
    collected = int(note.get('collected_count', 0) or 0)
    comment = int(note.get('comment_count', 0) or 0)
    share = int(note.get('share_count', 0) or 0)

    return {
        '序号': seq_num,
        '标题': sanitize(note.get('title', '') or '(无标题)'),
        '作者': sanitize(note.get('nickname', '')),
        '发布时间': date_str,
        '点赞数': liked,
        '收藏数': collected,
        '评论数': comment,
        '转发数': share,
        '总互动': liked + collected + comment + share,
        '标签': sanitize(tags),
        'IP属地': sanitize(note.get('ip_location', '')),
        '来源关键词': sanitize(source_kw or note.get('source_keyword', '')),
        '笔记链接': f"https://www.xiaohongshu.com/explore/{note.get('note_id', '')}",
        '笔记ID': note.get('note_id', ''),
    }


def find_latest_json(keyword_hint=""):
    pattern = os.path.join(PROJECT_DIR, "data", "xhs", "json", "search_contents_*.json")
    files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    if not files:
        return None
    return files[0]


def find_second_latest_json(latest_path):
    """找第二新的 JSON（第一轮采集的品牌词 vs 第二轮的睡眠词）"""
    pattern = os.path.join(PROJECT_DIR, "data", "xhs", "json", "search_contents_*.json")
    files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    if len(files) < 2:
        return None
    for f in files:
        if f != latest_path:
            return f
    return None


def load_xhs_data():
    """加载小红书采集数据，按 source_keyword 拆分品牌词 vs 睡眠词"""
    latest = find_latest_json()
    if not latest:
        print("未找到任何采集 JSON!")
        return [], []

    with open(latest, 'r', encoding='utf-8') as f:
        all_notes = json.load(f)

    # 按 source_keyword 拆分
    brand_notes = [n for n in all_notes if n.get('source_keyword', '') == '悟昕']
    sleep_notes = [n for n in all_notes if n.get('source_keyword', '') != '悟昕']

    # 如果最新 JSON 只有单关键词（旧模式），回退到双文件查找
    if not brand_notes or not sleep_notes:
        second = find_second_latest_json(latest)
        if second:
            with open(second, 'r', encoding='utf-8') as f:
                second_notes = json.load(f)
            for n in second_notes:
                if n.get('source_keyword', '') == '悟昕':
                    brand_notes.append(n)
                else:
                    sleep_notes.append(n)

    print(f"拆分结果: 品牌 {len(brand_notes)} 条, 睡眠 {len(sleep_notes)} 条")
    return brand_notes, sleep_notes


def filter_hot_notes(notes):
    seen_ids = set()
    filtered = []
    for note in notes:
        liked = int(note.get('liked_count', 0) or 0)
        nickname = (note.get('nickname', '') or '').lower().strip()
        note_id = note.get('note_id', '')
        if note_id in seen_ids:
            continue
        if liked < HOT_LIKE_THRESHOLD:
            continue
        if any(brand in nickname for brand in BRAND_ACCOUNTS):
            continue
        seen_ids.add(note_id)
        filtered.append(note)
    return filtered


def write_xhs_sheet(wb, sheet_name, notes, header_color="7C5DFA"):
    # 去重：同一 note_id 只保留一条（取点赞最高的版本）
    seen_ids = {}
    for n in notes:
        nid = n.get('note_id', '')
        if nid not in seen_ids:
            seen_ids[nid] = n
        else:
            # 保留互动数据更完整的
            existing_score = int(seen_ids[nid].get('liked_count', 0) or 0)
            new_score = int(n.get('liked_count', 0) or 0)
            if new_score > existing_score:
                seen_ids[nid] = n
    deduped = list(seen_ids.values())

    notes_sorted = sorted(deduped, key=lambda x: int(x.get('liked_count', 0) or 0), reverse=True)
    data = [normalize_note(n, i + 1) for i, n in enumerate(notes_sorted)]

    ws = wb.create_sheet(title=sheet_name)
    if not data:
        ws.cell(row=1, column=1, value="无数据")
        return 0

    headers = list(data[0].keys())
    write_styled_sheet(ws, data, headers, header_color)
    return len(data)


def write_styled_sheet(ws, data, headers, header_color):
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
            cell.alignment = Alignment(vertical="center", wrap_text=(key in ('标题', '摘要')))

    col_widths = {
        '序号': 6, '标题': 45, '作者': 15, '发布时间': 18,
        '点赞数': 10, '收藏数': 10, '评论数': 10, '转发数': 10,
        '总互动': 10, '标签': 30, 'IP属地': 10, '来源关键词': 12,
        '笔记链接': 50, '笔记ID': 25, '平台': 12, '摘要': 50,
        '互动数据': 18, '链接': 55, '日期': 14, '社区/来源': 20,
        '相关性评分': 12, '发布者': 18, '互动分': 10, '相关性': 8,
        '趋势点评': 55,
    }
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_widths.get(header, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def parse_last30days_output(text):
    """解析 last30days compact 输出，提取结构化条目"""
    items = []
    current_platform = ""
    current_title = ""
    current_score = 0

    lines = text.split('\n')
    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # 匹配 evidence cluster 标题行: ### 1. Title (score N, M items, sources: X)
        cluster_match = re.match(r'^###\s+\d+\.\s+(.+?)\s+\(score\s+(\d+)', line)
        if cluster_match:
            current_title = cluster_match.group(1).strip()
            current_score = int(cluster_match.group(2))
            i += 1
            continue

        # 匹配条目行: "1. [reddit] Title" 或 "1. [youtube] Title"
        item_match = re.match(r'^\d+\.\s+\[(\w+)\]\s+(.+)', line)
        if item_match:
            current_platform = item_match.group(1).capitalize()
            if current_platform == "Hn":
                current_platform = "Hacker News"
            # 替换 cluster title 为实际条目标题
            current_title = item_match.group(2).strip()
            i += 1
            continue

        # 匹配详情行: "- 2026-07-27 | r/GarminWatches | [186pts, 71cmt] | score:38"
        detail_match = re.match(r'^-\s+(\d{4}-\d{2}-\d{2})\s*\|\s*(.+?)\s*\|\s*\[([^\]]*)\]', line)
        if detail_match:
            date_str = detail_match.group(1)
            community = detail_match.group(2).strip()
            engagement = detail_match.group(3).strip()

            # 找 URL（下一行或下下行）
            url = ""
            for j in range(i+1, min(i+4, len(lines))):
                url_match = re.search(r'URL:\s*\[?(https?://\S+?)\]?', lines[j])
                if url_match:
                    url = url_match.group(1).rstrip(')')
                    break

            items.append({
                '序号': len(items) + 1,
                '平台': current_platform,
                '标题': current_title,
                '日期': date_str,
                '社区/来源': community,
                '互动数据': engagement,
                '相关性评分': current_score,
                '链接': url,
            })
            i += 1
            continue

        i += 1

    return items


def extract_author_from_url(url, source):
    """从 URL 提取作者/来源账号"""
    import re as _re
    url = url or ''
    source = (source or '').lower()
    try:
        if source == 'x' or source == 'twitter':
            # https://x.com/USERNAME/status/...
            m = _re.search(r'x\.com/(\w+)/status', url) or _re.search(r'twitter\.com/(\w+)/status', url)
            return f"@{m.group(1)}" if m else ''
        elif source == 'reddit':
            # https://www.reddit.com/r/SUBREDDIT/comments/...
            m = _re.search(r'/r/(\w+)/', url)
            return f"r/{m.group(1)}" if m else ''
        elif source == 'github':
            # https://github.com/ORG/REPO/issues/...
            m = _re.match(r'https?://github\.com/([^/]+)/([^/]+)/', url)
            return f"{m.group(1)}/{m.group(2)}" if m else ''
        elif source == 'youtube':
            return 'YouTube频道'
        elif source == 'hackernews':
            return 'HN用户'
    except Exception:
        pass
    return ''


def generate_trend_comment(r, platform, author):
    """根据条目内容生成一句话趋势点评"""
    title = (r.get('title', '') or '').lower()
    summary = (r.get('summary', '') or '').lower()
    text = f"{title} {summary}"
    eng = r.get('engagement', {})

    # 按主题分类点评
    if 'sleepagotchi' in text:
        return 'Sleepagotchi 游戏化睡眠追踪引热议，从数据收集转向习惯养成，用户讨论其向AI健康平台转型的可能性'
    if 'somnoa' in text:
        return 'Somnoa AI：端侧（local-first）睡眠追踪器，OLED屏+YAMNet边缘AI，正处于14天开发挑战中'
    if 'orion sleep' in text:
        return 'Orion Sleep System 用户差评退货：传感器单侧不准、客服响应慢，床垫式睡眠监测产品需解决可靠性问题'
    if 'galaxy watch' in text and 'sleep' in text:
        return '三星 Galaxy Watch9 促销，主打AI睡眠教练功能，大厂智能手表睡眠赛道竞争加剧'
    if 'wearable' in text and 'accurate' in text or 'sleep lab' in text:
        return '高价值对比帖：14项睡眠实验室研究横向对比Apple/Oura/Fitbit/Whoop等设备准确度，可作品类参考'
    if 'fitbit' in text and 'sleep' in text:
        return 'Fitbit Air 睡眠追踪遭用户吐槽准确度差，消费级手环睡眠监测信任度待提升'
    if 'merging wearable' in text or 'personalized sleep' in text:
        return 'AI+可穿戴数据融合做个性化睡眠方案被讨论，方向获认可但落地效果待验证'
    if 'bigger idea' in text or 'where the conversation' in text:
        return 'AI+健康赛道趋势讨论：睡眠追踪正从单一功能向AI健康平台演进'
    if 'advanced sleep phase' in text:
        return '睡眠相位综合征医学数据整理，专业医疗级睡眠知识库建设'
    if 'morning waiting' in text or 'stalesync' in text:
        return 'FloorLamp/Allos项目：睡眠数据同步时序问题讨论，技术向'
    return ''


def write_global_trend_sheet(wb, l30d_output_file):
    """将 last30days JSON 输出写入 Sheet 3"""
    # 优先用含 X 的版本，回退到普通版本
    json_file = "/tmp/last30days_with_x.json"
    if not os.path.exists(json_file):
        json_file = "/tmp/last30days_output.json"

    if not os.path.exists(json_file):
        # 回退到 compact text
        if l30d_output_file and os.path.exists(l30d_output_file):
            with open(l30d_output_file, 'r', encoding='utf-8') as f:
                text = f.read()
            items = parse_last30days_output(text)
            ws = wb.create_sheet(title="全球趋势")
            if items:
                headers = list(items[0].keys())
                write_styled_sheet(ws, items, headers, "E8654A")
                return len(items)
        ws = wb.create_sheet(title="全球趋势")
        ws.cell(row=1, column=1, value="last30days 采集失败，无数据")
        return 0

    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    results = data.get('results', [])

    # 过滤：relevance_score <= 0 直接丢弃（不相关内容）
    IRRELEVANT_KEYWORDS = ['anime', '动漫', '漫剧', 'multisub', '新番', '国漫',
                           'AI Law', 'ai-law', '二次元', '仙域', '劍宗']
    def is_irrelevant(r):
        rel = r.get('relevance_score', 0) or 0
        if rel <= 0:
            return True
        title = (r.get('title', '') or '').lower()
        summary = (r.get('summary', '') or '').lower()
        text = f"{title} {summary}"
        for kw in IRRELEVANT_KEYWORDS:
            if kw.lower() in text:
                return True
        return False

    filtered_results = [r for r in results if not is_irrelevant(r)]
    print(f"全球趋势过滤: {len(results)} -> {len(filtered_results)} 条 (去除 relevance<=0 及不相关)")

    # 转为表格行
    items = []
    for i, r in enumerate(filtered_results):
        eng = r.get('engagement', {})
        score = 0
        # 统一互动数据格式
        if isinstance(eng, dict):
            if 'likes' in eng or 'replies' in eng or 'reposts' in eng:
                likes = eng.get('likes', 0) or 0
                replies = eng.get('replies', 0) or 0
                reposts = eng.get('reposts', 0) or 0
                score = likes + replies + reposts
                eng_str = f"赞:{likes}, 评:{replies}, 转:{reposts}"
            elif 'score' in eng or 'points' in eng:
                score = eng.get('score') or eng.get('points') or 0
                comments = eng.get('num_comments') or eng.get('comments') or 0
                views = eng.get('views') or eng.get('view_count') or 0
                eng_str = f"score:{score}, cmt:{comments}"
                if views:
                    eng_str += f", views:{views}"
            else:
                score = sum(v for v in eng.values() if isinstance(v, (int, float)))
                eng_str = str(eng)
        else:
            eng_str = str(eng)

        source = r.get('source', '')
        platform_map = {
            'reddit': 'Reddit', 'youtube': 'YouTube', 'hackernews': 'Hacker News',
            'github': 'GitHub', 'hn': 'Hacker News', 'web': 'Web',
            'x': 'X/Twitter', 'twitter': 'X/Twitter',
        }
        platform = platform_map.get(source.lower(), source.capitalize() if source else "Unknown")

        # 提取作者/来源
        author = extract_author_from_url(r.get('url', ''), source)
        # 生成趋势点评
        comment = generate_trend_comment(r, platform, author)

        items.append({
            '序号': i + 1,
            '平台': platform,
            '发布者': author,
            '标题': sanitize(r.get('title', '')[:200]),
            '日期': r.get('published_at', '')[:10] if r.get('published_at') else '',
            '互动数据': eng_str,
            '互动分': score,
            '相关性': round(r.get('relevance_score', 0) or 0, 2),
            '趋势点评': comment,
            '摘要': sanitize((r.get('summary', '') or '')[:300]),
            '链接': r.get('url', ''),
        })

    # 按互动分降序
    items.sort(key=lambda x: x.get('互动分', 0), reverse=True)
    for i, item in enumerate(items):
        item['序号'] = i + 1

    ws = wb.create_sheet(title="全球趋势")
    if items:
        headers = list(items[0].keys())
        write_styled_sheet(ws, items, headers, "E8654A")
    else:
        ws.cell(row=1, column=1, value="无数据")

    return len(items)


def main():
    print("=" * 60)
    print("合并三 Sheet Excel 报告")
    print("=" * 60)

    # 加载小红书数据
    brand_notes, sleep_notes = load_xhs_data()
    print(f"品牌笔记: {len(brand_notes)} 条")
    print(f"睡眠笔记(原始): {len(sleep_notes)} 条")

    # 筛选爆款
    hot_notes = filter_hot_notes(sleep_notes)
    print(f"睡眠爆款(点赞>{HOT_LIKE_THRESHOLD}): {len(hot_notes)} 条")

    # 加载 last30days 数据
    l30d_file = "/tmp/last30days_output.txt"
    if os.path.exists(l30d_file):
        print(f"last30days 输出: {l30d_file}")
    else:
        print("警告: last30days 输出文件不存在")
        l30d_file = None

    # 创建 Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: 品牌笔记
    count1 = write_xhs_sheet(wb, "悟昕品牌笔记", brand_notes, "7C5DFA")
    print(f"Sheet 1 '悟昕品牌笔记': {count1} 条")

    # Sheet 2: 睡眠爆款
    count2 = write_xhs_sheet(wb, "睡眠爆款笔记", hot_notes, "E8654A")
    print(f"Sheet 2 '睡眠爆款笔记': {count2} 条")

    # Sheet 3: 全球趋势
    count3 = write_global_trend_sheet(wb, l30d_file)
    print(f"Sheet 3 '全球趋势': {count3} 条")

    # 保存
    today_str = datetime.now().strftime("%Y%m%d")
    output_filename = f"悟昕小红书笔记数据_MediaCrawler_{today_str}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, output_filename)

    if os.path.exists(output_path):
        os.remove(output_path)

    wb.save(output_path)
    print(f"\n✅ Excel 已保存: {output_path}")
    print("=" * 60)

    return output_path


if __name__ == "__main__":
    main()
