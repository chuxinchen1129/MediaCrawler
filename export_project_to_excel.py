#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将无醇啤酒小红书数据分析项目档案导出为Excel表格
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 创建Excel文件
wb = Workbook()
wb.remove(wb.active)  # 删除默认工作表

# 定义样式
header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
title_font = Font(name='微软雅黑', size=14, bold=True, color='000000')
subtitle_font = Font(name='微软雅黑', size=11, bold=True, color='4472C4')
border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ==================== 1. 基本信息 ====================
ws1 = wb.create_sheet('基本信息')
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 40

ws1['A1'] = '项目档案：无醇啤酒小红书数据分析'
ws1['A1'].font = title_font
ws1.merge_cells('A1:B1')
ws1['A1'].alignment = center_align

basic_info = [
    ['项目名称', '无醇啤酒小红书用户评论数据分析'],
    ['完成日期', '2026-01-26'],
    ['数据来源', '小红书平台'],
    ['搜索关键词', '无醇啤酒'],
    ['数据规模', '100条笔记 + 2,011条评论'],
    ['项目状态', '✅ 已完成']
]

row = 3
for item in basic_info:
    ws1.cell(row=row, column=1, value=item[0])
    ws1.cell(row=row, column=2, value=item[1])
    ws1.cell(row=row, column=1).font = Font(name='微软雅黑', size=11, bold=True)
    ws1.cell(row=row, column=2).font = Font(name='微软雅黑', size=11)
    row += 1

# ==================== 2. 项目目标 ====================
ws2 = wb.create_sheet('项目目标')
ws2.column_dimensions['A'].width = 50
ws2.column_dimensions['B'].width = 15

ws2['A1'] = '项目目标'
ws2['A1'].font = title_font
ws2.merge_cells('A1:B1')
ws2['A1'].alignment = center_align

goals = [
    ['分析消费者对无醇啤酒的购买意愿', 1],
    ['识别真实的消费场景', 2],
    ['评估品牌提及率和竞争格局', 3],
    ['构建用户画像', 4],
    ['提供产品优化和市场策略建议', 5]
]

row = 3
for goal in goals:
    ws2.cell(row=row, column=1, value=goal[0])
    ws2.cell(row=row, column=2, value=goal[1])
    ws2.cell(row=row, column=1).font = Font(name='微软雅黑', size=11)
    ws2.cell(row=row, column=2).font = Font(name='微软雅黑', size=11)
    ws2.cell(row=row, column=2).alignment = center_align
    row += 1

# ==================== 3. 核心发现 ====================
ws3 = wb.create_sheet('核心发现')
ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 25
ws3.column_dimensions['C'].width = 15

ws3['A1'] = '核心发现'
ws3['A1'].font = title_font
ws3.merge_cells('A1:C1')
ws3['A1'].alignment = center_align

# 消费者意愿分布
row = 3
ws3.cell(row=row, column=1, value='消费者意愿分布')
ws3.cell(row=row, column=1).font = subtitle_font
ws3.merge_cells(f'A{row}:C{row}')
row += 1

headers = ['类别', '占比', '备注']
for col, header in enumerate(headers, 1):
    cell = ws3.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border
row += 1

willingness_data = [
    ['愿意尝试', '78.6%', ''],
    ['不愿意尝试', '21.4%', '']
]

for data in willingness_data:
    for col, value in enumerate(data, 1):
        cell = ws3.cell(row=row, column=col, value=value)
        cell.font = Font(name='微软雅黑', size=11)
        cell.alignment = center_align if col > 1 else left_align
        cell.border = border
    row += 2

# 关键驱动因素
ws3.cell(row=row, column=1, value='愿意尝试的原因')
ws3.cell(row=row, column=1).font = subtitle_font
ws3.merge_cells(f'A{row}:C{row}')
row += 1

headers = ['原因', '占比', '排名']
for col, header in enumerate(headers, 1):
    cell = ws3.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border
row += 1

willing_reasons = [
    ['口感好', '50.7%', '1'],
    ['社交/聚会', '11.1%', '2'],
    ['健身/减肥', '9.0%', '3'],
    ['好奇尝鲜', '8.3%', '4'],
    ['替代传统啤酒', '6.5%', '5']
]

for data in willing_reasons:
    for col, value in enumerate(data, 1):
        cell = ws3.cell(row=row, column=col, value=value)
        cell.font = Font(name='微软雅黑', size=11)
        cell.alignment = center_align if col > 1 else left_align
        cell.border = border
    row += 1

row += 1
ws3.cell(row=row, column=1, value='不愿意尝试的原因')
ws3.cell(row=row, column=1).font = subtitle_font
ws3.merge_cells(f'A{row}:C{row}')
row += 1

headers = ['原因', '占比', '排名']
for col, header in enumerate(headers, 1):
    cell = ws3.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border
row += 1

unwilling_reasons = [
    ['口感不好', '48.3%', '1'],
    ['失望体验', '32.3%', '2'],
    ['价格/性价比', '10.5%', '3']
]

for data in unwilling_reasons:
    for col, value in enumerate(data, 1):
        cell = ws3.cell(row=row, column=col, value=value)
        cell.font = Font(name='微软雅黑', size=11)
        cell.alignment = center_align if col > 1 else left_align
        cell.border = border
    row += 1

# ==================== 4. 消费场景分析 ====================
ws4 = wb.create_sheet('消费场景分析')
ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 15

ws4['A1'] = '高频消费场景 TOP 8'
ws4['A1'].font = title_font
ws4.merge_cells('A1:C1')
ws4['A1'].alignment = center_align

row = 3
headers = ['排名', '场景', '提及次数']
for col, header in enumerate(headers, 1):
    cell = ws4.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border
row += 1

scenarios = [
    ['1', '聚会', '82'],
    ['2', '开车', '49'],
    ['3', '健身', '39'],
    ['4', '夏天', '35'],
    ['5', '宵夜', '32'],
    ['6', '火锅', '28'],
    ['7', '居家', '24'],
    ['8', '减肥', '22']
]

for data in scenarios:
    for col, value in enumerate(data, 1):
        cell = ws4.cell(row=row, column=col, value=value)
        cell.font = Font(name='微软雅黑', size=11)
        cell.alignment = center_align if col == 1 else left_align
        cell.border = border
    row += 1

# ==================== 5. 品牌分析 ====================
ws5 = wb.create_sheet('品牌分析')
ws5.column_dimensions['A'].width = 20
ws5.column_dimensions['B'].width = 30
ws5.column_dimensions['C'].width = 15
ws5.column_dimensions['D'].width = 40

ws5['A1'] = '品牌提及率分析'
ws5['A1'].font = title_font
ws5.merge_cells('A1:D1')
ws5['A1'].alignment = center_align

row = 3
headers = ['品牌', '提及率', '提及次数', '特点']
for col, header in enumerate(headers, 1):
    cell = ws5.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border
row += 1

brands = [
    ['新零', '26.8%', '98', '宣传多但用户讨论少'],
    ['青岛无醇', '13.4%', '49', '有"怪味"问题但整体正面72.5%'],
    ['其他无醇', '57.7%', '211', '其他品牌汇总']
]

for data in brands:
    for col, value in enumerate(data, 1):
        cell = ws5.cell(row=row, column=col, value=value)
        cell.font = Font(name='微软雅黑', size=11)
        cell.alignment = center_align if col in [1, 2, 3] else left_align
        cell.border = border
    row += 1

# ==================== 6. 用户画像 ====================
ws6 = wb.create_sheet('用户画像')
ws6.column_dimensions['A'].width = 20
ws6.column_dimensions['B'].width = 30
ws6.column_dimensions['C'].width = 15
ws6.column_dimensions['D'].width = 40

ws6['A1'] = '用户画像 - 四大类型'
ws6['A1'].font = title_font
ws6.merge_cells('A1:D1')
ws6['A1'].alignment = center_align

row = 3
headers = ['类型', '占比', '特征', '描述']
for col, header in enumerate(headers, 1):
    cell = ws6.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border
row += 1

personas = [
    ['社交型', '24%', '注重社交氛围', '聚会场景驱动，强调融入感'],
    ['实用型', '28%', '刚性需求', '开车、孕妇等需要0酒精的群体'],
    ['健康型', '22%', '健康导向', '健身、减肥、养生人群'],
    ['体验型', '26%', '尝鲜心理', '好奇尝鲜，愿意尝试新品']
]

for data in personas:
    for col, value in enumerate(data, 1):
        cell = ws6.cell(row=row, column=col, value=value)
        cell.font = Font(name='微软雅黑', size=11)
        cell.alignment = center_align if col in [1, 2, 3] else left_align
        cell.border = border
    row += 1

# ==================== 7. 高赞评论 ====================
ws7 = wb.create_sheet('高赞评论TOP3')
ws7.column_dimensions['A'].width = 10
ws7.column_dimensions['B'].width = 50
ws7.column_dimensions['C'].width = 15

ws7['A1'] = '高赞评论 TOP 3'
ws7['A1'].font = title_font
ws7.merge_cells('A1:C1')
ws7['A1'].alignment = center_align

row = 3
headers = ['排名', '评论内容', '点赞数']
for col, header in enumerate(headers, 1):
    cell = ws7.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border
row += 1

top_comments = [
    ['1', '痛风不语，只是一味发作', '8,404'],
    ['2', '人不能直接喝尿', '3,851'],
    ['3', '想喝有酒精没酒味的', '1,262']
]

for data in top_comments:
    for col, value in enumerate(data, 1):
        cell = ws7.cell(row=row, column=col, value=value)
        cell.font = Font(name='微软雅黑', size=11)
        cell.alignment = center_align if col in [1, 3] else left_align
        cell.border = border
    row += 1

# ==================== 8. 交付成果 ====================
ws8 = wb.create_sheet('交付成果')
ws8.column_dimensions['A'].width = 25
ws8.column_dimensions['B'].width = 50

ws8['A1'] = '项目交付成果清单'
ws8['A1'].font = title_font
ws8.merge_cells('A1:B1')
ws8['A1'].alignment = center_align

row = 3
deliverables = [
    # 数据文件
    ['数据文件', ''],
    ['原始数据', 'xhs_search_20260126_162338.xlsx'],
    ['清洗后数据', 'xhs_search_20260126_162338_cleaned.xlsx'],
    ['', '包含笔记数据、评论数据、热度TOP20等'],
    ['', ''],
    # 可视化图表
    ['可视化图表', ''],
    ['数据分析报告', 'data_analysis_report.png - 品牌提及率和热门话题分析'],
    ['消费场景分析', 'consumption_scenarios_analysis.png - 消费场景分析'],
    ['消费者洞察', 'consumer_insights_analysis.png - 消费者洞察'],
    ['新零品牌分析', 'xinling_brand_analysis.png - 新零品牌分析'],
    ['消费者意愿分析', 'consumer_intent_analysis.png - 消费者意愿分析'],
    ['', ''],
    # PPT演示文稿
    ['PPT演示文稿', ''],
    ['文件名', '无醇啤酒消费者意愿与场景语义分析.pptx'],
    ['页数', '10页'],
    ['配色', 'Teal & Coral (#5EA8A7, #FE4447)'],
    ['', ''],
    # 文档
    ['文档', ''],
    ['数据清洗说明', '数据清洗说明.md - 完整的数据清洗报告'],
    ['字体配置说明', '词云图使用配置.md - Matplotlib中文字体配置']
]

for item in deliverables:
    ws8.cell(row=row, column=1, value=item[0])
    ws8.cell(row=row, column=2, value=item[1])
    if item[0] and not item[1]:  # 分类标题
        ws8.cell(row=row, column=1).font = subtitle_font
        ws8.cell(row=row, column=1).alignment = left_align
    else:
        ws8.cell(row=row, column=1).font = Font(name='微软雅黑', size=11, bold=True)
        ws8.cell(row=row, column=2).font = Font(name='微软雅黑', size=11)
        ws8.cell(row=row, column=2).alignment = left_align
    row += 1

# ==================== 9. 关键建议 ====================
ws9 = wb.create_sheet('关键建议')
ws9.column_dimensions['A'].width = 20
ws9.column_dimensions['B'].width = 50
ws9.column_dimensions['C'].width = 15

ws9['A1'] = '关键建议'
ws9['A1'].font = title_font
ws9.merge_cells('A1:C1')
ws9['A1'].alignment = center_align

row = 3
ws9.cell(row=row, column=1, value='类别')
ws9.cell(row=row, column=2, value='建议内容')
ws9.cell(row=row, column=3, value='优先级')
ws9.row_dimensions[row].height = 25
for col in range(1, 4):
    cell = ws9.cell(row=row, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border
row += 1

recommendations = [
    ['产品优化', '解决口感问题 - 50.7%愿意因为口感，48.3%不愿意因为口感', '高'],
    ['产品优化', '消除"怪味" - 青岛无醇的主要问题', '高'],
    ['产品优化', '降低尝试门槛 - 小包装、试饮装', '高'],
    ['', '', ''],
    ['场景营销', '聚焦三大场景：聚会、开车、健身', '高'],
    ['场景营销', '强化"不扫兴+0负担"卖点', '中'],
    ['场景营销', '夏季、宵夜、火锅场景推广', '中'],
    ['', '', ''],
    ['品牌策略', '新零：从"声量驱动"转向"口碑驱动"', '中'],
    ['品牌策略', '青岛：优化口感，发挥品牌信任度优势', '中'],
    ['品牌策略', '增加UGC内容激励', '中'],
    ['', '', ''],
    ['精准触达', '社交型：强调氛围融入', '中'],
    ['精准触达', '实用型：强调安全可靠', '中'],
    ['精准触达', '健康型：强调低卡无负担', '中'],
    ['精准触达', '体验型：强调新奇体验', '中']
]

for data in recommendations:
    for col, value in enumerate(data, 1):
        if value:  # 只写入非空单元格
            cell = ws9.cell(row=row, column=col, value=value)
            cell.font = Font(name='微软雅黑', size=11)
            cell.alignment = left_align if col == 2 else center_align
            if value:  # 有内容时添加边框
                cell.border = border
    if any(data):  # 如果这一行有内容
        ws9.row_dimensions[row].height = 30
    row += 1

# ==================== 10. 后续研究方向 ====================
ws10 = wb.create_sheet('后续研究方向')
ws10.column_dimensions['A'].width = 10
ws10.column_dimensions['B'].width = 60

ws10['A1'] = '后续研究方向'
ws10['A1'].font = title_font
ws10.merge_cells('A1:B1')
ws10['A1'].alignment = center_align

row = 3
research_directions = [
    ['1', '竞品对比分析（新零 vs 青岛vs 进口品牌）'],
    ['2', '价格敏感度分析'],
    ['3', '地域分布分析'],
    ['4', '时间趋势分析（季节性波动）'],
    ['5', '用户画像深化（年龄、性别、职业）'],
    ['6', '扩大样本量到500-1000条笔记'],
    ['7', '多平台对比（微博、抖音、B站）']
]

for item in research_directions:
    ws10.cell(row=row, column=1, value=item[0])
    ws10.cell(row=row, column=2, value=item[1])
    ws10.cell(row=row, column=1).font = Font(name='微软雅黑', size=11, bold=True)
    ws10.cell(row=row, column=1).alignment = center_align
    ws10.cell(row=row, column=2).font = Font(name='微软雅黑', size=11)
    ws10.cell(row=row, column=2).alignment = left_align
    row += 1

# ==================== 11. 技术方案 ====================
ws11 = wb.create_sheet('技术方案')
ws11.column_dimensions['A'].width = 20
ws11.column_dimensions['B'].width = 50

ws11['A1'] = '技术方案与工具'
ws11['A1'].font = title_font
ws11.merge_cells('A1:B1')
ws11['A1'].alignment = center_align

row = 3
ws11.cell(row=row, column=1, value='类别')
ws11.cell(row=row, column=2, value='说明')
ws11.cell(row=row, column=1).font = subtitle_font
ws11.cell(row=row, column=2).font = subtitle_font
row += 1

tech_info = [
    ['数据清洗', '时间戳转换、新增热度字段、内容长度计算'],
    ['中文字体配置', '使用STZHONGS.TTF字体解决matplotlib中文乱码'],
    ['分析工具', 'Python 3.x, pandas, matplotlib, jieba, PptxGenJS'],
    ['数据可视化', 'matplotlib生成各类统计图表'],
    ['PPT生成', 'PptxGenJS + html2pptx'],
    ['数据处理', 'pandas数据清洗与转换'],
    ['中文分词', 'jieba分词与关键词提取']
]

for item in tech_info:
    ws11.cell(row=row, column=1, value=item[0])
    ws11.cell(row=row, column=2, value=item[1])
    ws11.cell(row=row, column=1).font = Font(name='微软雅黑', size=11, bold=True)
    ws11.cell(row=row, column=2).font = Font(name='微软雅黑', size=11)
    ws11.cell(row=row, column=2).alignment = left_align
    row += 1

# ==================== 12. 技术问题解决 ====================
ws12 = wb.create_sheet('技术问题解决')
ws12.column_dimensions['A'].width = 20
ws12.column_dimensions['B'].width = 30
ws12.column_dimensions['C'].width = 15
ws12.column_dimensions['D'].width = 40

ws12['A1'] = '技术问题与解决方案'
ws12['A1'].font = title_font
ws12.merge_cells('A1:D1')
ws12['A1'].alignment = center_align

row = 3
headers = ['问题', '现象', '状态', '解决方案']
for col, header in enumerate(headers, 1):
    cell = ws12.cell(row=row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border
row += 1

issues = [
    ['中文乱码', 'matplotlib图表中文显示为方框', '✅ 已解决', '配置STZHONGS.TTF字体'],
    ['数据格式不统一', '时间戳格式混乱，数值字段为字符串', '✅ 已解决', '编写数据清洗脚本，批量转换'],
    ['html2pptx验证严格', '渐变不支持、文本元素不能有背景、内容溢出', '✅ 已解决', '改用纯色、div包装、调整padding']
]

for data in issues:
    for col, value in enumerate(data, 1):
        cell = ws12.cell(row=row, column=col, value=value)
        cell.font = Font(name='微软雅黑', size=11)
        cell.alignment = center_align if col in [1, 3] else left_align
        cell.border = border
    row += 1

# ==================== 保存文件 ====================
output_file = '/Users/echo/MediaCrawler/无醇啤酒小红书数据分析项目档案.xlsx'
wb.save(output_file)
print(f"✅ Excel文件已生成: {output_file}")
print(f"📊 共包含 {len(wb.worksheets)} 个工作表:")
for idx, sheet in enumerate(wb.worksheets, 1):
    print(f"   {idx}. {sheet.title}")
